"""Project BOQ Excel importer + budget summary APIs.

Reads the SBI_BOQ_Template_v2.xlsx (sheets: Rate Analysis, Stages, BOQ)
and creates/updates:
  - Site Resource masters (prices)
  - Site Activity masters (coefficients + computed 5-category rates)
  - Project stage Tasks (is_group=1) if missing
  - Project BOQ document with stage-wise lines and category split
"""

import io

import frappe
import openpyxl
from frappe.utils import flt
from frappe.utils.file_manager import get_file

RA_SHEET = "Rate Analysis"
ST_SHEET = "Stages"
BOQ_SHEET = "BOQ"

RA_COL_START = 11      # column K
RA_ACT_START = 8       # first activity row
RA_ACT_END = 45
ST_ROW_START = 4
ST_ROW_END = 15
BOQ_ROW_START = 3
BOQ_ROW_END = 202


def _get_file_bytes(file_url):
    """Return raw bytes of an uploaded File given its file_url."""
    _fname, content = get_file(file_url)
    if isinstance(content, str):
        content = content.encode("utf-8")
    return content


def _load_wb(file_url):
    content = _get_file_bytes(file_url)
    return openpyxl.load_workbook(io.BytesIO(content), data_only=True)


def _import_resources(ws):
    created, updated = 0, 0
    col = RA_COL_START
    while True:
        name = ws.cell(row=4, column=col).value
        if not name:
            break
        name = str(name).strip()
        if name and not name.lower().startswith("spare"):
            cat = str(ws.cell(row=3, column=col).value or "Other").strip()
            unit = str(ws.cell(row=5, column=col).value or "").strip()
            rate = flt(ws.cell(row=6, column=col).value)
            if cat not in ("Labour", "Material", "Equipment", "Subcontract", "Other"):
                cat = "Other"
            if frappe.db.exists("Site Resource", name):
                frappe.db.set_value("Site Resource", name,
                                    {"category": cat, "unit": unit, "rate": rate})
                updated += 1
            else:
                frappe.get_doc({"doctype": "Site Resource",
                                "resource_name": name, "category": cat,
                                "unit": unit, "rate": rate}).insert()
                created += 1
        col += 1
    return created, updated, col - 1


def _import_activities(ws, last_col):
    created, updated = 0, 0
    for r in range(RA_ACT_START, RA_ACT_END + 1):
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        name = str(name).strip()
        unit = str(ws.cell(row=r, column=3).value or "").strip()
        if frappe.db.exists("Site Activity", name):
            doc = frappe.get_doc("Site Activity", name)
            updated += 1
        else:
            doc = frappe.new_doc("Site Activity")
            doc.activity_name = name
            created += 1
        doc.output_unit = unit
        doc.set("resources", [])
        for col in range(RA_COL_START, last_col + 1):
            res = ws.cell(row=4, column=col).value
            coef = ws.cell(row=r, column=col).value
            if not res or not coef:
                continue
            res = str(res).strip()
            if res.lower().startswith("spare"):
                continue
            if not frappe.db.exists("Site Resource", res):
                continue
            doc.append("resources", {"resource": res, "coefficient": flt(coef)})
        doc.save()
    return created, updated


def _import_stages(ws, project):
    """Ensure a group Task exists per stage. Returns {stage_no: label}."""
    stage_map = {}
    created = 0
    for r in range(ST_ROW_START, ST_ROW_END + 1):
        sno = ws.cell(row=r, column=2).value
        desc = ws.cell(row=r, column=3).value
        if not sno:
            continue
        sno = str(sno).strip()
        label = sno + " - " + str(desc or "").strip() if desc else sno
        stage_map[sno] = label
        exists = frappe.db.exists("Task",
                                  {"project": project, "subject": label})
        if not exists:
            frappe.get_doc({"doctype": "Task", "subject": label,
                            "project": project, "is_group": 1,
                            "status": "Open"}).insert()
            created += 1
    return stage_map, created


@frappe.whitelist()
def import_boq_excel(project, file_url, stage=None):
    """Parse template and build/update the Project BOQ.

    stage=None  -> full import (replaces all lines)
    stage given -> stage-wise import: only that stage's rows are read
                   from the file and only that stage's existing lines
                   are replaced; other stages stay untouched.
    Accepts 'Stage 2' or the full 'Stage 2 - <description>' label."""
    frappe.only_for(("System Manager", "Projects Manager"))
    if not frappe.db.exists("Project", project):
        frappe.throw("Project not found: " + project)

    wb = _load_wb(file_url)
    for sheet in (RA_SHEET, ST_SHEET, BOQ_SHEET):
        if sheet not in wb.sheetnames:
            frappe.throw("Sheet missing in file: " + sheet)

    ra = wb[RA_SHEET]
    res_created, res_updated, last_col = _import_resources(ra)
    act_created, act_updated = _import_activities(ra, last_col)
    stage_map, stages_created = _import_stages(wb[ST_SHEET], project)

    target_label = None
    if stage:
        stage = str(stage).strip()
        target_label = stage_map.get(stage, stage)

    existing = frappe.db.get_value("Project BOQ", {"project": project}, "name")
    if existing:
        boq = frappe.get_doc("Project BOQ", existing)
    else:
        boq = frappe.new_doc("Project BOQ")
        boq.project = project
    boq.source_file = file_url
    if target_label:
        keep = [dict(stage=i.stage, activity=i.activity, qty=i.qty,
                     rate=i.rate)
                for i in (boq.items or []) if i.stage != target_label]
        boq.set("items", [])
        for k in keep:
            boq.append("items", k)
    else:
        boq.set("items", [])

    ws = wb[BOQ_SHEET]
    lines = 0
    for r in range(BOQ_ROW_START, BOQ_ROW_END + 1):
        row_stage = ws.cell(row=r, column=1).value
        activity = ws.cell(row=r, column=2).value
        qty = flt(ws.cell(row=r, column=4).value)
        if not row_stage or not activity or not qty:
            continue
        row_stage = str(row_stage).strip()
        activity = str(activity).strip()
        if not frappe.db.exists("Site Activity", activity):
            continue
        label = stage_map.get(row_stage, row_stage)
        if target_label and label != target_label:
            continue
        boq.append("items", {"stage": label, "activity": activity, "qty": qty})
        lines += 1

    boq.save()
    frappe.db.commit()
    return {"boq": boq.name, "lines": lines,
            "stage_scope": target_label or "ALL",
            "resources_created": res_created, "resources_updated": res_updated,
            "activities_created": act_created, "activities_updated": act_updated,
            "stages_created": stages_created,
            "total_amount": boq.total_amount}


@frappe.whitelist()
def get_boq_summary(project):
    """Office/owner view: stage x category budget matrix WITH amounts."""
    frappe.only_for(("System Manager", "Projects Manager", "Accounts Manager"))
    name = frappe.db.get_value("Project BOQ", {"project": project}, "name")
    if not name:
        return {"stages": [], "totals": {}}
    boq = frappe.get_doc("Project BOQ", name)
    stages = {}
    for it in boq.items:
        s = stages.setdefault(it.stage, {"stage": it.stage, "labour": 0,
                                         "material": 0, "equipment": 0,
                                         "subcontract": 0, "other": 0,
                                         "total": 0})
        s["labour"] += flt(it.labour_amount)
        s["material"] += flt(it.material_amount)
        s["equipment"] += flt(it.equipment_amount)
        s["subcontract"] += flt(it.subcontract_amount)
        s["other"] += flt(it.other_amount)
        s["total"] += flt(it.amount)
    return {"stages": list(stages.values()),
            "totals": {"labour": boq.total_labour,
                       "material": boq.total_material,
                       "equipment": boq.total_equipment,
                       "subcontract": boq.total_subcontract,
                       "other": boq.total_other,
                       "total": boq.total_amount}}


@frappe.whitelist()
def get_site_stages(project):
    """Site app view: stage list + activities + qty ONLY. No amounts,
    no rates, no budgets (Q3 principle - site staff never see money)."""
    name = frappe.db.get_value("Project BOQ", {"project": project}, "name")
    if not name:
        return []
    items = frappe.get_all("Project BOQ Item",
                           filters={"parent": name},
                           fields=["stage", "activity", "unit", "qty"],
                           order_by="idx")
    stages = {}
    for it in items:
        stages.setdefault(it.stage, []).append(
            {"activity": it.activity, "unit": it.unit, "qty": it.qty})
    return [{"stage": k, "activities": v} for k, v in stages.items()]
