# Copyright (c) 2026, Velmaska and contributors
"""Builds the PO annexure workbook for a fabrication purchase order.

One workbook, several linked sheets, so a fabricator can trace a PO
category down to individual sections, cut lengths, assemblies and
shipping pieces. Returns raw xlsx bytes (attached to the PO in ERPNext).
"""

import io


def build_annexure(data, cats, po_no="PO"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    NAVY, BLUE, LT, GREY = "1F3864", "1F4E78", "DDEBF7", "F2F2F2"
    H = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    B = Font(bold=True, name="Calibri", size=10)
    N = Font(name="Calibri", size=10)
    T = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    thin = Border(*[Side("thin", color="B0B0B0")] * 4)
    ctr = Alignment("center", "center", wrap_text=True)
    lft = Alignment("left", "center", wrap_text=True)
    rgt = Alignment("right", "center")
    yel = PatternFill("solid", fgColor="FFF2CC")

    def title(ws, ncol, sub):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        c = ws.cell(1, 1, "SHIV BHARAT INFRASTRUCTURES")
        c.fill = PatternFill("solid", fgColor=NAVY); c.font = T; c.alignment = ctr
        ws.row_dimensions[1].height = 22
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        c = ws.cell(2, 1, f"Fabrication PO Annexure  |  {sub}  |  PO Ref: {po_no}")
        c.fill = PatternFill("solid", fgColor=BLUE); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = lft
        return 4

    def head(ws, r, cols):
        for i, (h, w) in enumerate(cols, 1):
            c = ws.cell(r, i, h); c.fill = PatternFill("solid", fgColor=BLUE)
            c.font = H; c.alignment = ctr; c.border = thin
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[r].height = 24
        return r + 1

    def row(ws, r, vals, al, shade=False):
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v); c.font = N; c.border = thin
            c.alignment = al[i - 1]
            if shade:
                c.fill = PatternFill("solid", fgColor=GREY)
            if isinstance(v, float):
                c.number_format = "#,##0.00"
        return r + 1

    # ---------------- Sheet 1: PO (category, rate/kg) ----------------
    ws = wb.create_sheet("PO by Category")
    r = title(ws, 5, "Purchase Order - Category-wise")
    r = head(ws, r, [("#", 5), ("Category", 26), ("Item Code", 16),
                     ("Weight (Kg)", 16), ("Rate/Kg (fill)", 16)])
    hdr = r - 1
    first = r
    for i, c in enumerate(cats, 1):
        ws.cell(r, 1, i).alignment = ctr; ws.cell(r, 1).font = N; ws.cell(r, 1).border = thin
        ws.cell(r, 2, c["label"]).alignment = lft; ws.cell(r, 2).font = N; ws.cell(r, 2).border = thin
        ws.cell(r, 3, c["item_code"]).alignment = ctr; ws.cell(r, 3).font = N; ws.cell(r, 3).border = thin
        wc = ws.cell(r, 4, round(c["weight_kg"], 2)); wc.number_format = "#,##0.00"; wc.alignment = rgt; wc.font = N; wc.border = thin
        rc = ws.cell(r, 5); rc.fill = yel; rc.border = thin
        r += 1
    tot = ws.cell(r, 2, "TOTAL"); tot.font = B
    ws.cell(r, 4, f"=SUM(D{first}:D{r-1})").font = B; ws.cell(r, 4).number_format = "#,##0.00"
    for i in range(1, 6):
        ws.cell(r, i).fill = PatternFill("solid", fgColor=LT); ws.cell(r, i).border = thin
    ws.sheet_view.showGridLines = False; ws.freeze_panes = f"A{hdr+1}"

    # ---------------- Sheet 2: Category -> Sections (linked) ----------------
    ws = wb.create_sheet("Category Detail")
    r = title(ws, 5, "Category -> Sections")
    r = head(ws, r, [("Category", 24), ("Section", 18), ("Grade / Spec", 30),
                     ("Weight (Kg)", 14), ("Nos", 10)])
    hdr = r - 1
    for c in cats:
        span_start = r
        for s in c["sections"]:
            r = row(ws, r, [c["label"], s["section"], s.get("grade", ""),
                            round(s["kg"], 2), s.get("nos", "")],
                    [lft, lft, lft, rgt, ctr], shade=(cats.index(c) % 2 == 1))
        # merge the category cell down its section rows
        if r - span_start > 1:
            ws.merge_cells(start_row=span_start, start_column=1, end_row=r - 1, end_column=1)
            ws.cell(span_start, 1).alignment = Alignment("left", "top", wrap_text=True)
            ws.cell(span_start, 1).font = B
    ws.sheet_view.showGridLines = False; ws.freeze_panes = f"A{hdr+1}"

    # ---------------- Sheet 3: Cutting List (SECD) ----------------
    ws = wb.create_sheet("Cutting List")
    r = title(ws, 6, "Cutting List (Section Details)")
    r = head(ws, r, [("Sl", 5), ("Part Mark", 14), ("Section", 18), ("Grade", 20),
                     ("Cut Length mm", 14), ("Member Type", 20)])
    hdr = r - 1; sl = 0
    for cr in data["cutting"]:
        sl += 1
        r = row(ws, r, [sl, cr["part"], cr["profile"], cr.get("grade", ""),
                        round(float(cr["length"]), 1) if cr["length"] else 0,
                        str(cr.get("member", "")).replace("_", " ").title()],
                [ctr, lft, lft, lft, ctr, lft], shade=(sl % 2 == 0))
    ws.sheet_view.showGridLines = False; ws.freeze_panes = f"A{hdr+1}"
    ws.auto_filter.ref = f"A{hdr}:F{r-1}"

    # ---------------- Sheet 4: Shipping List (SHIPG) ----------------
    ws = wb.create_sheet("Shipping List")
    r = title(ws, 5, "Shipping List (Dispatch)")
    r = head(ws, r, [("Sl", 5), ("Assembly Mark", 16), ("External Size mm", 22),
                     ("Total Wt kg", 14), ("Member Type", 22)])
    hdr = r - 1; sl = 0
    for sp in data["shipping"]:
        sl += 1
        r = row(ws, r, [sl, sp["asm"], sp.get("ext", ""),
                        round(float(sp["tot_wt"]), 2) if sp["tot_wt"] else 0,
                        str(sp.get("member", "")).replace("_", " ").title()],
                [ctr, lft, lft, rgt, lft], shade=(sl % 2 == 0))
    ws.sheet_view.showGridLines = False; ws.freeze_panes = f"A{hdr+1}"
    ws.auto_filter.ref = f"A{hdr}:E{r-1}"

    # ---------------- Sheet 5: Bolt List (BOLT) ----------------
    ws = wb.create_sheet("Bolt List")
    r = title(ws, 5, "Bolt List (bought-out)")
    r = head(ws, r, [("Sl", 5), ("Material Code", 28), ("Qty Nos", 12),
                     ("Total Wt kg", 14), ("Remark", 24)])
    hdr = r - 1; sl = 0
    for b in data["bolt"]:
        sl += 1
        r = row(ws, r, [sl, b["code"], int(b["qty"]),
                        round(float(b["tot_wt"]), 2) if b["tot_wt"] else 0, b.get("remark", "")],
                [ctr, lft, ctr, rgt, lft], shade=(sl % 2 == 0))
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
