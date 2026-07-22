# Copyright (c) 2026, Velmaska and contributors
"""PO annexure workbook + HTML for a fabrication purchase order.

Per-PO (single category): PO by Category, Category Detail, Cutting List
(filtered to that category's section family), and — for the Bolts PO — a
Bolt List. Shipping is dispatch-level (crosses categories) so it appears
only in the full workbook attached to the import document.
"""

import io

from sbi_projects.sbi_projects.steel_bom.parser import _family_of


def build_annexure(data, cats, po_no="PO"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    NAVY, BLUE, LT, GREY = "1F3864", "1F4E78", "DDEBF7", "F2F2F2"
    H = Font(bold=True, color="FFFFFF", size=10)
    B = Font(bold=True, size=10)
    N = Font(size=10)
    T = Font(bold=True, color="FFFFFF", size=14)
    thin = Border(*[Side("thin", color="B0B0B0")] * 4)
    ctr = Alignment("center", "center", wrap_text=True)
    lft = Alignment("left", "center", wrap_text=True)
    rgt = Alignment("right", "center")
    yel = PatternFill("solid", fgColor="FFF2CC")

    single = len(cats) == 1
    scope = cats[0]["label"] if single else "All Categories"
    wanted_keys = {c.get("key") for c in cats}

    def title(ws, ncol, sub):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        c = ws.cell(1, 1, "SHIV BHARAT INFRASTRUCTURES")
        c.fill = PatternFill("solid", fgColor=NAVY); c.font = T; c.alignment = ctr
        ws.row_dimensions[1].height = 22
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        c = ws.cell(2, 1, f"Fabrication PO Annexure  |  {sub}  |  {scope}  |  Ref: {po_no}")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.font = Font(bold=True, color="FFFFFF", size=10); c.alignment = lft
        return 4

    def head(ws, r, cols):
        for i, (h, w) in enumerate(cols, 1):
            c = ws.cell(r, i, h); c.fill = PatternFill("solid", fgColor=BLUE)
            c.font = H; c.alignment = ctr; c.border = thin
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[r].height = 24
        return r + 1

    def row(ws, r, vals, al, shade=False):
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v); c.font = N; c.border = thin; c.alignment = al[i - 1]
            if shade:
                c.fill = PatternFill("solid", fgColor=GREY)
            if isinstance(v, float):
                c.number_format = "#,##0.00"
        return r + 1

    # ---- Sheet 1: PO by Category ----
    ws = wb.create_sheet("PO by Category")
    r = title(ws, 5, "Purchase Order")
    r = head(ws, r, [("#", 5), ("Category", 26), ("Item Code", 16),
                     ("Weight (Kg)", 16), ("Rate/Kg (fill)", 16)])
    hdr = r - 1; first = r
    for i, c in enumerate(cats, 1):
        row(ws, r, [i, c["label"], c["item_code"], round(c["weight_kg"], 2), ""],
            [ctr, lft, ctr, rgt, ctr])
        ws.cell(r, 5).fill = yel
        r += 1
    ws.cell(r, 2, "TOTAL").font = B
    tc = ws.cell(r, 4, f"=SUM(D{first}:D{r-1})"); tc.font = B; tc.number_format = "#,##0.00"
    for i in range(1, 6):
        ws.cell(r, i).fill = PatternFill("solid", fgColor=LT); ws.cell(r, i).border = thin
    ws.sheet_view.showGridLines = False; ws.freeze_panes = f"A{hdr+1}"

    # ---- Sheet 2: Category Detail (sections) ----
    ws = wb.create_sheet("Category Detail")
    r = title(ws, 5, "Sections")
    r = head(ws, r, [("Category", 24), ("Section", 18), ("Grade / Spec", 30),
                     ("Weight (Kg)", 14), ("Nos", 10)])
    hdr = r - 1
    for ci, c in enumerate(cats):
        span = r
        for s in c["sections"]:
            row(ws, r, [c["label"], s["section"], s.get("grade", ""),
                        round(s["kg"], 2), s.get("nos", "")],
                [lft, lft, lft, rgt, ctr], shade=(ci % 2 == 1))
            r += 1
        if r - span > 1:
            ws.merge_cells(start_row=span, start_column=1, end_row=r - 1, end_column=1)
            ws.cell(span, 1).alignment = Alignment("left", "top", wrap_text=True)
            ws.cell(span, 1).font = B
    ws.sheet_view.showGridLines = False; ws.freeze_panes = f"A{hdr+1}"

    # ---- Sheet 3: Cutting List (filtered to these categories) ----
    ws = wb.create_sheet("Cutting List")
    r = title(ws, 6, "Cutting List (Section Details)")
    r = head(ws, r, [("Sl", 5), ("Part Mark", 14), ("Section", 18), ("Grade", 20),
                     ("Cut Length mm", 14), ("Member Type", 20)])
    hdr = r - 1; sl = 0
    for cr in data["cutting"]:
        if _family_of(cr["profile"]) not in wanted_keys:
            continue
        sl += 1
        row(ws, r, [sl, cr["part"], cr["profile"], cr.get("grade", ""),
                    round(float(cr["length"]), 1) if cr["length"] else 0,
                    str(cr.get("member", "")).replace("_", " ").title()],
            [ctr, lft, lft, lft, ctr, lft], shade=(sl % 2 == 0))
        r += 1
    ws.sheet_view.showGridLines = False; ws.freeze_panes = f"A{hdr+1}"
    if sl:
        ws.auto_filter.ref = f"A{hdr}:F{r-1}"

    # ---- Sheet 4: Shipping List (full workbook only) ----
    if not single:
        ws = wb.create_sheet("Shipping List")
        r = title(ws, 5, "Shipping List (Dispatch)")
        r = head(ws, r, [("Sl", 5), ("Assembly Mark", 16), ("External Size mm", 22),
                         ("Total Wt kg", 14), ("Member Type", 22)])
        hdr = r - 1; sl = 0
        for sp in data["shipping"]:
            sl += 1
            row(ws, r, [sl, sp["asm"], sp.get("ext", ""),
                        round(float(sp["tot_wt"]), 2) if sp["tot_wt"] else 0,
                        str(sp.get("member", "")).replace("_", " ").title()],
                [ctr, lft, lft, rgt, lft], shade=(sl % 2 == 0))
            r += 1
        ws.sheet_view.showGridLines = False; ws.freeze_panes = f"A{hdr+1}"

    # ---- Sheet 5: Bolt List (only when bolts category present) ----
    if any(c.get("key") == "BOLTS" for c in cats):
        ws = wb.create_sheet("Bolt List")
        r = title(ws, 5, "Bolt List (bought-out)")
        r = head(ws, r, [("Sl", 5), ("Material Code", 28), ("Qty Nos", 12),
                         ("Total Wt kg", 14), ("Remark", 24)])
        hdr = r - 1; sl = 0
        for b in data["bolt"]:
            sl += 1
            row(ws, r, [sl, b["code"], int(b["qty"]),
                        round(float(b["tot_wt"]), 2) if b["tot_wt"] else 0, b.get("remark", "")],
                [ctr, lft, ctr, rgt, lft], shade=(sl % 2 == 0))
            r += 1
        ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ======================================================================
# HTML annexure (used by the PO print format AND the PDF attachment)
# ======================================================================
def build_annexure_html(cat, data, po_no="PO"):
    label = cat["label"]
    wanted = cat.get("key")

    th = "background:#1F4E78;color:#fff;padding:4px 6px;text-align:left;border:1px solid #bbb;"
    td = "padding:3px 6px;border:1px solid #ccc;"
    tbl = "border-collapse:collapse;width:100%;font-size:11px;margin:6px 0;"

    secs = "".join(
        f"<tr><td style='{td}'>{s['section']}</td>"
        f"<td style='{td}'>{s.get('grade','')}</td>"
        f"<td style='{td};text-align:right'>{s['kg']:,.2f}</td>"
        f"<td style='{td};text-align:right'>{s.get('nos','') or ''}</td></tr>"
        for s in cat["sections"]
    )

    cut_rows = [c for c in data.get("cutting", []) if _family_of(c["profile"]) == wanted][:80]
    cuts = "".join(
        f"<tr><td style='{td}'>{c['part']}</td>"
        f"<td style='{td}'>{c['profile']}</td>"
        f"<td style='{td};text-align:right'>{(c['length'] or 0):,.0f}</td>"
        f"<td style='{td}'>{str(c.get('member','')).replace('_',' ').title()}</td></tr>"
        for c in cut_rows
    )
    cut_note = "" if len(cut_rows) < 80 else "<p style='font-size:10px'>First 80 cutting lines; full list in the attached annexure.</p>"

    cut_block = ""
    if cuts:
        cut_block = (
            "<h5 style='margin:10px 0 4px'>Cutting reference (part / section / cut length mm / member)</h5>"
            f"<table style='{tbl}'>"
            f"<tr><th style='{th}'>Part Mark</th><th style='{th}'>Section</th>"
            f"<th style='{th}'>Cut Length</th><th style='{th}'>Member</th></tr>"
            f"{cuts}</table>{cut_note}"
        )

    return f"""
<div style="margin-top:14px">
  <h4 style="color:#BE1E2D;margin:8px 0 4px">Annexure — {label} (Ref {po_no})</h4>
  <table style="{tbl}">
    <tr><th style="{th}">Section</th><th style="{th}">Grade / Spec</th>
        <th style="{th}">Weight (Kg)</th><th style="{th}">Nos</th></tr>
    {secs}
  </table>
  {cut_block}
</div>
"""


# ======================================================================
# PDF annexure (single category) - rendered via frappe's wkhtmltopdf
# ======================================================================
def build_annexure_pdf(cat, data, po_no="PO"):
    from frappe.utils.pdf import get_pdf

    body = build_annexure_html(cat, data, po_no=po_no)
    html = (
        "<div style='font-family:sans-serif'>"
        "<h2 style='color:#1F3864;margin:0'>SHIV BHARAT INFRASTRUCTURES</h2>"
        f"<p style='margin:2px 0 10px;color:#555'>Fabrication PO Annexure "
        f"&nbsp;|&nbsp; <b>{cat['label']}</b> &nbsp;|&nbsp; Ref: {po_no} "
        f"&nbsp;|&nbsp; Total: {cat['weight_kg']:,.2f} Kg</p>"
        f"{body}</div>"
    )
    return get_pdf(html)
