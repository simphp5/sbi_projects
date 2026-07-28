# Copyright (c) 2026, Velmaska and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import flt


class EstimationSheetBOQ(Document):
	def validate(self):
		self.compute_amounts()
		self.compute_totals()
		self.check_line_stages()

	def compute_amounts(self):
		for row in (self.lines or []):
			row.qty = flt(row.qty)
			row.rate = flt(row.rate)
			row.amount = flt(row.qty) * flt(row.rate)

	def compute_totals(self):
		base = sum(flt(r.amount) for r in (self.lines or []))
		self.base_total = base

		markup = 0.0
		for pct in (self.site_establishment_pct, self.contingency_pct,
		            self.overhead_pct, self.profit_pct, self.incentive_pct):
			markup += base * (flt(pct) / 100.0)

		self.markup_total = markup
		self.grand_total = base + markup

		area = flt(self.built_up_area)
		self.rate_per_sft = (base + markup) / area if area else 0
		self.stage_count = len(self.stages or [])

	@frappe.whitelist()
	def load_stages(self, source=None, replace=1):
		"""Fill the stage list from a Payment Terms Template or the linked Sales Order.

		Stages are the payment terms -- keeping one source of truth means the BOQ,
		the quotation and the eventual invoices all speak the same sequence.
		Typing stages in by hand stays possible; this just saves the typing.
		"""
		source = source or ("Sales Order" if self.sales_order else "Payment Terms Template")

		if source == "Sales Order":
			if not self.sales_order:
				frappe.throw(_("Link a Sales Order first."))
			rows = frappe.get_all(
				"Payment Schedule",
				filters={"parent": self.sales_order, "parenttype": "Sales Order"},
				fields=["payment_term", "description", "invoice_portion", "idx"],
				order_by="idx asc",
			)
		else:
			if not self.payment_terms_template:
				frappe.throw(_("Select a Payment Terms Template first."))
			rows = frappe.get_all(
				"Payment Terms Template Detail",
				filters={"parent": self.payment_terms_template},
				fields=["payment_term", "description", "invoice_portion", "idx"],
				order_by="idx asc",
			)

		if not rows:
			frappe.throw(_("No payment terms found in {0}.").format(frappe.bold(source)))

		if replace in (1, "1", True, "true"):
			self.stages = []

		existing = {(r.stage_name or "").strip().lower() for r in (self.stages or [])}
		added = 0
		for r in rows:
			name = (r.payment_term or r.description or "").strip()
			if not name or name.lower() in existing:
				continue
			self.append("stages", {
				"stage_name": name,
				"description": r.description,
				"invoice_portion": flt(r.invoice_portion),
				"source": source,
			})
			existing.add(name.lower())
			added += 1

		self.compute_totals()
		self.save(ignore_permissions=True)
		frappe.db.commit()
		return {"added": added, "source": source}

	@frappe.whitelist()
	def stage_options(self):
		"""Stage names defined on this sheet, for the line-level picker."""
		return [r.stage_name for r in (self.stages or []) if r.stage_name]

	def check_line_stages(self):
		"""Flag BOQ lines whose stage is not one of the defined stages.

		This warns rather than blocks -- a sheet is often drafted before the
		payment terms are settled, and stopping the save would be unhelpful.
		"""
		defined = {(r.stage_name or "").strip().lower() for r in (self.stages or [])}
		if not defined:
			return
		stray = sorted({
			(l.stage or "").strip() for l in (self.lines or [])
			if (l.stage or "").strip() and (l.stage or "").strip().lower() not in defined
		})
		if stray:
			frappe.msgprint(
				_("These line stages are not in the stage list: {0}").format(
					frappe.bold(", ".join(stray))
				),
				title=_("Stage mismatch"),
				indicator="orange",
			)

	@frappe.whitelist()
	def recalculate_rates(self):
		"""Refresh WA-source line rates from the Rate Card. Manual lines untouched."""
		card = self.rate_card
		if not card:
			from sbi_projects.peb_estimation.doctype.rate_card.rate_card import RateCard
			card = RateCard.get_default()
		if not card:
			frappe.throw(_("No Rate Card selected and no default Rate Card found."))
		updated = 0
		for row in (self.lines or []):
			if row.rate_source == "Manual":
				continue
			if not row.work_item:
				continue
			wi = frappe.get_cached_doc("Work Item", row.work_item)
			row.rate = wi.computed_rate(card)
			row.amount = flt(row.qty) * flt(row.rate)
			updated += 1
		self.compute_totals()
		self.save(ignore_permissions=True)
		return updated

	@frappe.whitelist()
	def import_lines(self, file_url, replace=0):
		"""Append (or replace) BOQ lines from an uploaded Excel/CSV file.

		Expected columns (header row, case-insensitive, order-independent):
			Stage, Category, Work Item, Description, UOM, Qty, Rate, Remarks

		Rules applied per row:
		  - Qty / Rate are coerced with flt, so blanks and text are safe.
		  - A rate supplied in the file marks the line Manual, so a later
		    "Recalculate Rates (WA)" will not overwrite the imported price.
		  - A blank rate leaves the line on WA, to be priced from the Rate Card.
		  - Unknown Work Item codes are reported, not silently dropped.
		"""
		if not file_url:
			frappe.throw(_("Attach a file first."))

		rows = _read_tabular(file_url)
		if len(rows) < 2:
			frappe.throw(_("The file has no data rows."))

		header = [str(c or "").strip().lower() for c in rows[0]]
		required = "qty"
		if required not in header:
			frappe.throw(
				_("Could not find a <b>Qty</b> column. Please use the downloaded template.")
			)

		def col(*names):
			for n in names:
				if n in header:
					return header.index(n)
			return None

		idx = {
			"stage": col("stage"),
			"category": col("category"),
			"work_item": col("work item", "work_item", "item"),
			"description": col("description", "particulars"),
			"uom": col("uom", "unit"),
			"qty": col("qty", "quantity"),
			"rate": col("rate"),
			"remarks": col("remarks", "remark"),
		}

		if replace in (1, "1", True, "true"):
			self.lines = []

		valid_items = set(
			frappe.get_all("Work Item", filters={"is_active": 1}, pluck="name")
		)

		added, unknown, skipped = 0, [], 0
		for r_no, raw in enumerate(rows[1:], start=2):
			def val(key):
				i = idx.get(key)
				if i is None or i >= len(raw):
					return None
				v = raw[i]
				return str(v).strip() if v is not None else None

			work_item = val("work_item")
			description = val("description")
			qty = flt(val("qty"))

			# a row with nothing usable is just blank spacing in the sheet
			if not work_item and not description and not qty:
				skipped += 1
				continue

			if work_item and work_item not in valid_items:
				unknown.append(_("Row {0}: {1}").format(r_no, work_item))
				work_item = None

			rate = flt(val("rate"))
			row = self.append("lines", {
				"stage": val("stage"),
				"category": val("category"),
				"work_item": work_item,
				"description": description,
				"uom": val("uom"),
				"qty": qty,
				"rate": rate,
				"rate_source": "Manual" if rate else "WA",
				"remarks": val("remarks"),
			})
			row.amount = flt(row.qty) * flt(row.rate)
			added += 1

		self.compute_totals()
		self.save(ignore_permissions=True)
		frappe.db.commit()

		return {"added": added, "skipped": skipped, "unknown": unknown}


def _read_tabular(file_url):
	"""Return a list of rows from an attached .xlsx / .xls / .csv file."""
	import os

	ext = os.path.splitext((file_url or "").split("?")[0])[1].lower()

	if ext == ".csv":
		from frappe.utils.csvutils import read_csv_content

		file_doc = frappe.get_doc("File", {"file_url": file_url})
		return read_csv_content(file_doc.get_content())

	if ext in (".xlsx", ".xls"):
		from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file

		return read_xlsx_file_from_attached_file(file_url=file_url)

	frappe.throw(_("Unsupported file type {0}. Use .xlsx or .csv.").format(ext or "?"))


@frappe.whitelist()
def download_boq_template():
	"""Stream an .xlsx template: BOQ Lines sheet + a reference sheet of Work Items."""
	from io import BytesIO

	import openpyxl
	from openpyxl.styles import Alignment, Font, PatternFill

	headers = ["Stage", "Category", "Work Item", "Description", "UOM", "Qty", "Rate", "Remarks"]

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "BOQ Lines"

	head_fill = PatternFill("solid", fgColor="BE1E2D")
	head_font = Font(color="FFFFFF", bold=True)
	for c, h in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=c, value=h)
		cell.fill = head_fill
		cell.font = head_font
		cell.alignment = Alignment(horizontal="center")

	for c, w in enumerate([14, 14, 20, 40, 10, 10, 12, 24], start=1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
	ws.freeze_panes = "A2"

	# two illustrative rows -- one WA-priced, one manually priced
	samples = [
		["Stage 1", "Civil", "CIV-RCC-M25", "RCC M25 for footings", "CUM", 25, "", "leave Rate blank to price from Rate Card"],
		["Stage 1", "Civil", "CIV-PCC-M10", "PCC bedding", "CUM", 12, 5460, "Rate filled = Manual, kept on Recalculate"],
	]
	for r, row in enumerate(samples, start=2):
		for c, v in enumerate(row, start=1):
			ws.cell(row=r, column=c, value=v)

	# reference sheet so the estimator can copy exact codes
	ref = wb.create_sheet("Work Items")
	for c, h in enumerate(["Work Item Code", "Name", "Category", "UOM"], start=1):
		cell = ref.cell(row=1, column=c, value=h)
		cell.fill = head_fill
		cell.font = head_font
	for c, w in enumerate([22, 38, 16, 10], start=1):
		ref.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
	ref.freeze_panes = "A2"

	items = frappe.get_all(
		"Work Item",
		filters={"is_active": 1},
		fields=["name", "work_item_name", "category", "uom"],
		order_by="category asc, name asc",
	)
	for r, it in enumerate(items, start=2):
		ref.cell(row=r, column=1, value=it.name)
		ref.cell(row=r, column=2, value=it.work_item_name)
		ref.cell(row=r, column=3, value=it.category)
		ref.cell(row=r, column=4, value=it.uom)

	buf = BytesIO()
	wb.save(buf)

	frappe.response["filename"] = "estimation_sheet_boq_template.xlsx"
	frappe.response["filecontent"] = buf.getvalue()
	frappe.response["type"] = "binary"

	# ------------------------------------------------------------------ #
	# Quotation
	# ------------------------------------------------------------------ #
	@frappe.whitelist()
	def create_quotation(self, group_by="Line", valid_days=30):
		"""Roll this BOQ into a Quotation for the client.

		Markup is absorbed into the line rates, so the quotation shows clean
		selling rates and the client never sees the internal build-up. The
		quotation total therefore equals this sheet's grand total exactly --
		the absorption factor is applied to rates and the last line carries any
		rounding remainder.

		group_by:
			Line     - one quotation row per BOQ line (full detail)
			Category - one row per category (Civil, Structural, ...)
			Stage    - one row per stage
		"""
		if self.quotation and frappe.db.exists("Quotation", self.quotation):
			frappe.throw(
				_("This BOQ already has a quotation: {0}").format(
					frappe.utils.get_link_to_form("Quotation", self.quotation)
				)
			)
		if not (self.lines or []):
			frappe.throw(_("Add BOQ lines first."))
		if not self.customer:
			frappe.throw(_("Set the Customer before creating a quotation."))

		self.compute_amounts()
		self.compute_totals()

		base = flt(self.base_total)
		if base <= 0:
			frappe.throw(_("The BOQ base total is zero, so there is nothing to quote."))

		# markup absorbed into rates
		factor = flt(self.grand_total) / base

		rows = self._quotation_rows(group_by, factor)
		if not rows:
			frappe.throw(_("No priced lines to quote."))

		quo = frappe.new_doc("Quotation")
		quo.quotation_to = "Customer"
		quo.party_name = self.customer
		quo.transaction_date = frappe.utils.today()
		quo.valid_till = frappe.utils.add_days(frappe.utils.today(), int(valid_days or 30))
		if self.payment_terms_template:
			quo.payment_terms_template = self.payment_terms_template

		for r in rows:
			quo.append("items", r)

		quo.insert(ignore_permissions=True)

		# nudge the total onto the BOQ grand total, absorbing rounding on the last row
		diff = flt(self.grand_total) - flt(quo.total)
		if abs(diff) >= 0.01 and quo.items:
			last = quo.items[-1]
			last.rate = flt(last.rate) + diff / flt(last.qty or 1)
			quo.save(ignore_permissions=True)

		self.db_set("quotation", quo.name)
		self.db_set("status", "Quoted")
		if self.building_enquiry:
			frappe.db.set_value("Building Enquiry", self.building_enquiry, {
				"quotation": quo.name,
				"status": "Quoted",
			})
		frappe.db.commit()

		return {"quotation": quo.name, "rows": len(rows), "total": flt(quo.total)}

	def _quotation_rows(self, group_by, factor):
		"""Build quotation item rows, with markup already folded into the rate."""
		if group_by == "Line":
			rows = []
			for l in (self.lines or []):
				if flt(l.amount) <= 0:
					continue
				item_code = self._item_for(l)
				rows.append({
					"item_code": item_code,
					"description": l.description or None,
					"qty": flt(l.qty) or 1,
					"uom": l.uom or None,
					"rate": flt(l.rate) * factor,
				})
			return rows

		# consolidated: one row per category or per stage, priced as a lump sum
		key_field = "category" if group_by == "Category" else "stage"
		buckets = {}
		for l in (self.lines or []):
			if flt(l.amount) <= 0:
				continue
			key = (getattr(l, key_field, None) or _("Other")).strip()
			buckets.setdefault(key, 0)
			buckets[key] += flt(l.amount)

		lump = _lump_sum_item()
		return [
			{
				"item_code": lump,
				"description": key,
				"qty": 1,
				"rate": amount * factor,
			}
			for key, amount in buckets.items()
		]

	def _item_for(self, line):
		"""Resolve the sales Item for a BOQ line, falling back to a lump-sum item."""
		if line.work_item:
			wi = frappe.get_doc("Work Item", line.work_item)
			return wi.get_or_create_item()
		return _lump_sum_item()


def _lump_sum_item():
	"""Generic sales item for consolidated or work-item-less rows."""
	code = "CONSTRUCTION-WORKS"
	if frappe.db.exists("Item", code):
		return code

	from sbi_projects.peb_estimation.doctype.work_item.work_item import (
		_construction_item_group,
	)

	frappe.get_doc({
		"doctype": "Item",
		"item_code": code,
		"item_name": "Construction Works",
		"description": "Construction works as per BOQ",
		"item_group": _construction_item_group(),
		"stock_uom": "Nos",
		"is_stock_item": 0,
		"is_sales_item": 1,
		"is_purchase_item": 0,
	}).insert(ignore_permissions=True)
	return code
